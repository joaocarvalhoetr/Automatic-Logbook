import openpyxl
from openpyxl import load_workbook
import csv
from datetime import datetime
import os
import shutil

MAX_FLIGHTS_PER_SHEET = 1500
TEMPLATE_SHEET_NAME = "Sheet0"
HEADER_ROW = 2

def load_excel(file_path, template_path):
    """Carregar o ficheiro Excel ou criar um novo a partir do template."""
    if not os.path.exists(file_path):
        shutil.copy(template_path, file_path)
    wb = load_workbook(file_path)
    return wb

def find_next_available_row(sheet):
    """Encontrar a próxima linha disponível na folha."""
    for row in range(HEADER_ROW + 1, HEADER_ROW + 1 + MAX_FLIGHTS_PER_SHEET):
        if not sheet.cell(row=row, column=2).value:  # Assumindo que a coluna 2 (B) está sempre preenchida
            return row
    return None

def load_aircraft_data(csv_file_path):
    """Carregar dados das aeronaves do ficheiro CSV."""
    aircraft_data = {}
    with open(csv_file_path, mode='r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            registration = row['Registration'].replace("-", "").upper()
            aircraft_data[registration] = {
                'model': row['Aircraft Type'],
                'variant': 'B737-800' if '737-800' in row['Aircraft Type'] else 'B737-8200' if 'MAX 8' in row['Aircraft Type'] else 'NA'
            }
    return aircraft_data

def get_all_flights(wb):
    """Extrair todos os voos do logbook."""
    flights = []
    for sheet in wb.worksheets:
        if sheet.title != TEMPLATE_SHEET_NAME:
            for row in range(HEADER_ROW + 1, HEADER_ROW + 1 + MAX_FLIGHTS_PER_SHEET):
                date_cell = sheet.cell(row=row, column=2).value
                departure_time_cell = sheet.cell(row=row, column=4).value
                if not date_cell or not departure_time_cell:
                    continue  # Ignore rows without a date or departure time
                flight_data = {
                    'date': date_cell,
                    'departure_airport': sheet.cell(row=row, column=3).value,
                    'departure_time': departure_time_cell,
                    'arrival_airport': sheet.cell(row=row, column=5).value,
                    'arrival_time': sheet.cell(row=row, column=6).value,
                    'aircraft_registration': sheet.cell(row=row, column=8).value,
                    'flight_time': sheet.cell(row=row, column=11).value,
                    'captain': sheet.cell(row=row, column=13).value,
                    'takeoffs_day': sheet.cell(row=row, column=14).value,
                    'takeoffs_night': sheet.cell(row=row, column=15).value,
                    'landings_day': sheet.cell(row=row, column=16).value,
                    'landings_night': sheet.cell(row=row, column=17).value,
                    'night_flight_time': sheet.cell(row=row, column=18).value,
                    'ifr_time': sheet.cell(row=row, column=19).value,
                }
                flights.append(flight_data)
    return flights

def add_flight_to_sheet(sheet, row, flight_data, aircraft_data):
    """Adicionar dados de voo à linha dada na folha."""
    registration_no_dash = flight_data['aircraft_registration'].replace("-", "").upper()
    aircraft_info = aircraft_data.get(registration_no_dash, {'model': 'NA', 'variant': 'NA'})

    # Desmesclar células na linha especificada
    for merged_cell in sheet.merged_cells.ranges:
        if merged_cell.min_row == row and merged_cell.max_row == row:
            sheet.unmerge_cells(str(merged_cell))

    sheet.cell(row=row, column=2, value=flight_data['date'])
    sheet.cell(row=row, column=3, value=flight_data['departure_airport'])
    sheet.cell(row=row, column=4, value=flight_data['departure_time'])
    sheet.cell(row=row, column=5, value=flight_data['arrival_airport'])
    sheet.cell(row=row, column=6, value=flight_data['arrival_time'])
    sheet.cell(row=row, column=7, value=aircraft_info['variant'])
    sheet.cell(row=row, column=8, value=flight_data['aircraft_registration'])
    sheet.cell(row=row, column=11, value=flight_data['flight_time'])
    sheet.cell(row=row, column=12, value=flight_data['flight_time'])
    sheet.cell(row=row, column=13, value=flight_data['captain'])
    sheet.cell(row=row, column=14, value=flight_data['takeoffs_day'])
    sheet.cell(row=row, column=15, value=flight_data['takeoffs_night'])
    sheet.cell(row=row, column=16, value=flight_data['landings_day'])
    sheet.cell(row=row, column=17, value=flight_data['landings_night'])
    sheet.cell(row=row, column=18, value=flight_data['night_flight_time'])
    sheet.cell(row=row, column=19, value=flight_data['ifr_time'])
    sheet.cell(row=row, column=21, value=flight_data['flight_time'])


def create_new_sheet(wb):
    """Criar uma nova folha a partir da template."""
    new_sheet_index = len(wb.sheetnames)
    new_sheet_name = f"Sheet{new_sheet_index}"
    new_sheet = wb.copy_worksheet(wb[TEMPLATE_SHEET_NAME])
    new_sheet.title = new_sheet_name
    return new_sheet

def save_excel(wb, file_path):
    """Salvar o ficheiro Excel."""
    wb.save(file_path)

def reorganize_logbook(file_path, csv_file_path):
    wb = load_excel(file_path, 'logbook_template.xlsx')
    aircraft_data = load_aircraft_data(csv_file_path)
    flights = get_all_flights(wb)

    # Ordenar voos por data e hora
    for flight in flights:
        flight['datetime'] = datetime.strptime(f"{flight['date']} {flight['departure_time']}", '%Y/%m/%d %H:%M')
    flights.sort(key=lambda x: x['datetime'])

    # Remover todas as folhas, exceto a template
    sheets_to_remove = [sheet for sheet in wb.worksheets if sheet.title != TEMPLATE_SHEET_NAME]
    for sheet in sheets_to_remove:
        wb.remove(sheet)

    # Adicionar os voos em novas folhas
    current_sheet = create_new_sheet(wb)
    current_row = HEADER_ROW + 1

    for flight in flights:
        if current_row > MAX_FLIGHTS_PER_SHEET + HEADER_ROW:
            current_sheet = create_new_sheet(wb)
            current_row = HEADER_ROW + 1

        add_flight_to_sheet(current_sheet, current_row, flight, aircraft_data)
        current_row += 1

    save_excel(wb, file_path)

def add_flights_to_excel(file_path, flights, csv_file_path):
    wb = load_excel(file_path, '_internal/logbook_template.xlsx')
    aircraft_data = load_aircraft_data(csv_file_path)
    current_sheet = None
    previous_sheet = None

    # Encontrar uma folha existente com linhas disponíveis ou criar uma nova folha
    for sheet in wb.worksheets:
        if sheet.title != TEMPLATE_SHEET_NAME:
            row = find_next_available_row(sheet)
            if row is not None:
                current_sheet = sheet
                break

    if current_sheet is None:
        current_sheet = create_new_sheet(wb)
    else:
        previous_sheet = current_sheet

    for flight in flights:
        row = find_next_available_row(current_sheet)
        if row is None:
            previous_sheet = current_sheet
            current_sheet = create_new_sheet(wb)
            row = find_next_available_row(current_sheet)
        add_flight_to_sheet(current_sheet, row, flight, aircraft_data)

    save_excel(wb, file_path)

